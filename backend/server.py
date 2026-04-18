from fastapi import FastAPI, APIRouter, Depends, HTTPException, status, Header, UploadFile, File, BackgroundTasks
from fastapi.responses import JSONResponse, Response
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from datetime import datetime, timezone
from typing import Optional, List
import json
import uuid
import base64

from models import (
    User, UserCreate, UserLogin, Token,
    Warehouse, WarehouseCreate, WarehouseUpdate,
    Supplier, SupplierCreate, SupplierUpdate,
    Product, ProductCreate, ProductUpdate,
    Invoice, InvoiceCreate, OCRRequest,
    Sale, SaleCreate,
    DashboardStats, StockAlert, FinancialReport, AuditLog,
    AlertConfig, AlertConfigCreate, Notification, NotificationCreate,
    Order, OrderCreate, OrderUpdate, CashFlowReport
)
from auth import hash_password, verify_password, create_token, decode_token
from audit import AuditLogger
from emergentintegrations.llm.chat import LlmChat, UserMessage, ImageContent
from email_service import send_email, build_stock_alert_email, build_invoice_pending_email, build_sale_completed_email
from report_export import generate_financial_pdf, generate_financial_excel
from nfe_parser import parse_nfe_xml

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

audit_logger = AuditLogger(db)

app = FastAPI()
api_router = APIRouter(prefix="/api")

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

async def get_current_user(authorization: Optional[str] = Header(None)) -> dict:
    if not authorization or not authorization.startswith('Bearer '):
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Not authenticated"
        )
    
    token = authorization.replace('Bearer ', '')
    payload = decode_token(token)
    
    if not payload:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Invalid or expired token"
        )
    
    return payload

async def require_role(user: dict, allowed_roles: List[str]):
    if user['role'] not in allowed_roles:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Insufficient permissions"
        )

@api_router.post("/auth/register", response_model=User)
async def register(user_data: UserCreate, current_user: dict = Depends(get_current_user)):
    await require_role(current_user, ["dev", "master"])
    
    existing = await db.users.find_one({"email": user_data.email}, {"_id": 0})
    if existing:
        raise HTTPException(status_code=400, detail="Email already registered")
    
    hashed_pw = hash_password(user_data.password)
    user_dict = user_data.model_dump(exclude={'password'})
    user_obj = User(**user_dict, created_at=datetime.now(timezone.utc))
    
    doc = user_obj.model_dump()
    doc['password_hash'] = hashed_pw
    doc['created_at'] = doc['created_at'].isoformat()
    
    await db.users.insert_one(doc)
    
    await audit_logger.log(
        current_user['user_id'],
        current_user['email'],
        "CREATE",
        "user",
        user_obj.id,
        {"email": user_obj.email, "role": user_obj.role}
    )
    
    return user_obj

@api_router.post("/auth/login", response_model=Token)
async def login(credentials: UserLogin):
    user_doc = await db.users.find_one({"email": credentials.email}, {"_id": 0})
    
    if not user_doc or not verify_password(credentials.password, user_doc['password_hash']):
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Invalid email or password"
        )
    
    if not user_doc.get('active', True):
        raise HTTPException(status_code=403, detail="User account is inactive")
    
    user_doc.pop('password_hash')
    user_doc['created_at'] = datetime.fromisoformat(user_doc['created_at'])
    user = User(**user_doc)
    
    token = create_token(user.id, user.email, user.role)
    
    return Token(access_token=token, user=user)

@api_router.get("/auth/me", response_model=User)
async def get_me(current_user: dict = Depends(get_current_user)):
    user_doc = await db.users.find_one({"id": current_user['user_id']}, {"_id": 0, "password_hash": 0})
    if not user_doc:
        raise HTTPException(status_code=404, detail="User not found")
    user_doc['created_at'] = datetime.fromisoformat(user_doc['created_at'])
    return User(**user_doc)

@api_router.get("/users", response_model=List[User])
async def get_users(current_user: dict = Depends(get_current_user)):
    await require_role(current_user, ["dev", "master"])
    
    users = await db.users.find({}, {"_id": 0, "password_hash": 0}).to_list(1000)
    for u in users:
        u['created_at'] = datetime.fromisoformat(u['created_at'])
    return [User(**u) for u in users]

@api_router.patch("/users/{user_id}")
async def update_user(user_id: str, updates: dict, current_user: dict = Depends(get_current_user)):
    await require_role(current_user, ["dev", "master"])
    
    if 'password' in updates:
        updates['password_hash'] = hash_password(updates.pop('password'))
    
    result = await db.users.update_one({"id": user_id}, {"$set": updates})
    
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    
    await audit_logger.log(
        current_user['user_id'],
        current_user['email'],
        "UPDATE",
        "user",
        user_id,
        updates
    )
    
    return {"message": "User updated successfully"}

@api_router.post("/warehouses", response_model=Warehouse)
async def create_warehouse(warehouse_data: WarehouseCreate, current_user: dict = Depends(get_current_user)):
    warehouse = Warehouse(
        **warehouse_data.model_dump(),
        created_at=datetime.now(timezone.utc),
        created_by=current_user['user_id']
    )
    
    doc = warehouse.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    
    await db.warehouses.insert_one(doc)
    
    await audit_logger.log(
        current_user['user_id'],
        current_user['email'],
        "CREATE",
        "warehouse",
        warehouse.id
    )
    
    return warehouse

@api_router.get("/warehouses", response_model=List[Warehouse])
async def get_warehouses(current_user: dict = Depends(get_current_user)):
    warehouses = await db.warehouses.find({}, {"_id": 0}).to_list(1000)
    for w in warehouses:
        w['created_at'] = datetime.fromisoformat(w['created_at'])
    return [Warehouse(**w) for w in warehouses]

@api_router.post("/suppliers", response_model=Supplier)
async def create_supplier(supplier_data: SupplierCreate, current_user: dict = Depends(get_current_user)):
    supplier = Supplier(
        **supplier_data.model_dump(),
        created_at=datetime.now(timezone.utc),
        created_by=current_user['user_id']
    )
    
    doc = supplier.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    
    await db.suppliers.insert_one(doc)
    
    await audit_logger.log(
        current_user['user_id'],
        current_user['email'],
        "CREATE",
        "supplier",
        supplier.id
    )
    
    return supplier

@api_router.get("/suppliers", response_model=List[Supplier])
async def get_suppliers(current_user: dict = Depends(get_current_user)):
    suppliers = await db.suppliers.find({}, {"_id": 0}).to_list(1000)
    for s in suppliers:
        s['created_at'] = datetime.fromisoformat(s['created_at'])
    return [Supplier(**s) for s in suppliers]

@api_router.post("/products", response_model=Product)
async def create_product(product_data: ProductCreate, current_user: dict = Depends(get_current_user)):
    existing = await db.products.find_one({"sku": product_data.sku}, {"_id": 0})
    if existing:
        raise HTTPException(status_code=400, detail="SKU already exists")
    
    product = Product(
        **product_data.model_dump(),
        created_at=datetime.now(timezone.utc),
        created_by=current_user['user_id']
    )
    
    doc = product.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    
    await db.products.insert_one(doc)
    
    await audit_logger.log(
        current_user['user_id'],
        current_user['email'],
        "CREATE",
        "product",
        product.id
    )
    
    return product

@api_router.get("/products", response_model=List[Product])
async def get_products(current_user: dict = Depends(get_current_user)):
    products = await db.products.find({}, {"_id": 0}).to_list(5000)
    for p in products:
        p['created_at'] = datetime.fromisoformat(p['created_at'])
    return [Product(**p) for p in products]

@api_router.get("/inventory")
async def get_inventory(current_user: dict = Depends(get_current_user)):
    inventory = await db.inventory.find({}, {"_id": 0}).to_list(5000)
    
    result = []
    for item in inventory:
        product = await db.products.find_one({"id": item['product_id']}, {"_id": 0})
        warehouse = await db.warehouses.find_one({"id": item['warehouse_id']}, {"_id": 0})
        
        if product and warehouse:
            result.append({
                "id": item['id'],
                "product_id": item['product_id'],
                "product_name": product['name'],
                "product_sku": product['sku'],
                "warehouse_id": item['warehouse_id'],
                "warehouse_name": warehouse['name'],
                "quantity": item['quantity'],
                "min_stock": product.get('min_stock', 0),
                "updated_at": item['updated_at']
            })
    
    return result

@api_router.post("/inventory/adjust")
async def adjust_inventory(
    product_id: str,
    warehouse_id: str,
    quantity: float,
    current_user: dict = Depends(get_current_user)
):
    existing = await db.inventory.find_one(
        {"product_id": product_id, "warehouse_id": warehouse_id},
        {"_id": 0}
    )
    
    if existing:
        new_qty = existing['quantity'] + quantity
        if new_qty < 0:
            new_qty = 0
        await db.inventory.update_one(
            {"id": existing['id']},
            {"$set": {"quantity": new_qty, "updated_at": datetime.now(timezone.utc).isoformat()}}
        )
    else:
        item = {
            "id": str(uuid.uuid4()),
            "product_id": product_id,
            "warehouse_id": warehouse_id,
            "quantity": quantity,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }
        await db.inventory.insert_one(item)
    
    await audit_logger.log(
        current_user['user_id'],
        current_user['email'],
        "ADJUST",
        "inventory",
        f"{product_id}:{warehouse_id}",
        {"quantity": quantity}
    )
    
    return {"message": "Inventory adjusted successfully"}

@api_router.post("/invoices", response_model=Invoice)
async def create_invoice(invoice_data: InvoiceCreate, current_user: dict = Depends(get_current_user)):
    invoice = Invoice(
        **invoice_data.model_dump(),
        created_at=datetime.now(timezone.utc),
        created_by=current_user['user_id']
    )
    
    doc = invoice.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    
    await db.invoices.insert_one(doc)
    
    await audit_logger.log(
        current_user['user_id'],
        current_user['email'],
        "CREATE",
        "invoice",
        invoice.id
    )
    
    return invoice

@api_router.get("/invoices", response_model=List[Invoice])
async def get_invoices(current_user: dict = Depends(get_current_user)):
    invoices = await db.invoices.find({}, {"_id": 0}).to_list(5000)
    for inv in invoices:
        inv['created_at'] = datetime.fromisoformat(inv['created_at'])
    return [Invoice(**inv) for inv in invoices]

@api_router.post("/invoices/ocr")
async def process_invoice_ocr(ocr_request: OCRRequest, current_user: dict = Depends(get_current_user)):
    try:
        api_key = os.environ.get('EMERGENT_LLM_KEY')
        if not api_key:
            raise HTTPException(status_code=500, detail="API key not configured")
        
        chat = LlmChat(
            api_key=api_key,
            session_id=f"ocr_{current_user['user_id']}_{datetime.now(timezone.utc).timestamp()}",
            system_message="You are an expert at extracting data from Brazilian fiscal invoices (NFe). Extract all relevant information in JSON format."
        ).with_model("openai", "gpt-4o")
        
        image_content = ImageContent(image_base64=ocr_request.image_base64)
        
        prompt_text = """Extract the following information from this Brazilian invoice image and return ONLY a valid JSON object with this structure:
{
  "invoice_number": "number",
  "supplier_name": "supplier name",
  "issue_date": "YYYY-MM-DD",
  "total_value": 0.0,
  "tax_value": 0.0,
  "items": [
    {
      "product_name": "name",
      "quantity": 0.0,
      "unit_price": 0.0,
      "total": 0.0
    }
  ]
}
Return ONLY the JSON, no explanations."""
        
        user_message = UserMessage(
            text=prompt_text,
            file_contents=[image_content]
        )
        
        response = await chat.send_message(user_message)
        
        response_text = response.strip()
        if response_text.startswith('```json'):
            response_text = response_text[7:]
        if response_text.startswith('```'):
            response_text = response_text[3:]
        if response_text.endswith('```'):
            response_text = response_text[:-3]
        response_text = response_text.strip()
        
        extracted_data = json.loads(response_text)
        
        return extracted_data
        
    except Exception as e:
        logger.error(f"OCR processing error: {e}")
        raise HTTPException(status_code=500, detail=f"OCR processing failed: {str(e)}")

@api_router.get("/dashboard/stats", response_model=DashboardStats)
async def get_dashboard_stats(current_user: dict = Depends(get_current_user)):
    total_products = await db.products.count_documents({"active": True})
    total_suppliers = await db.suppliers.count_documents({"active": True})
    total_warehouses = await db.warehouses.count_documents({"active": True})
    
    today = datetime.now(timezone.utc).date()
    today_start = datetime.combine(today, datetime.min.time()).replace(tzinfo=timezone.utc).isoformat()
    
    sales_today = await db.sales.find({
        "created_at": {"$gte": today_start},
        "status": "completed"
    }, {"_id": 0}).to_list(1000)
    total_sales_today = sum(s.get('total', 0) for s in sales_today)
    
    month_start = datetime.now(timezone.utc).replace(day=1, hour=0, minute=0, second=0).isoformat()
    sales_month = await db.sales.find({
        "created_at": {"$gte": month_start},
        "status": "completed"
    }, {"_id": 0}).to_list(5000)
    total_sales_month = sum(s.get('total', 0) for s in sales_month)
    
    inventory = await db.inventory.find({}, {"_id": 0}).to_list(5000)
    low_stock_count = 0
    for item in inventory:
        product = await db.products.find_one({"id": item['product_id']}, {"_id": 0})
        if product and item['quantity'] <= product.get('min_stock', 0):
            low_stock_count += 1
    
    pending_invoices = await db.invoices.count_documents({"status": "pending"})
    
    return DashboardStats(
        total_products=total_products,
        total_suppliers=total_suppliers,
        total_warehouses=total_warehouses,
        total_sales_today=total_sales_today,
        total_sales_month=total_sales_month,
        low_stock_alerts=low_stock_count,
        pending_invoices=pending_invoices
    )

@api_router.get("/dashboard/alerts", response_model=List[StockAlert])
async def get_stock_alerts(current_user: dict = Depends(get_current_user)):
    inventory = await db.inventory.find({}, {"_id": 0}).to_list(5000)
    alerts = []
    
    for item in inventory:
        product = await db.products.find_one({"id": item['product_id']}, {"_id": 0})
        warehouse = await db.warehouses.find_one({"id": item['warehouse_id']}, {"_id": 0})
        
        if product and warehouse and item['quantity'] <= product.get('min_stock', 0):
            alerts.append(StockAlert(
                product_id=item['product_id'],
                product_name=product['name'],
                warehouse_id=item['warehouse_id'],
                warehouse_name=warehouse['name'],
                current_quantity=item['quantity'],
                min_stock=product.get('min_stock', 0)
            ))
    
    return alerts

@api_router.get("/reports/financial")
async def get_financial_report(period: str, current_user: dict = Depends(get_current_user)):
    if period == "month":
        start_date = datetime.now(timezone.utc).replace(day=1, hour=0, minute=0, second=0).isoformat()
    else:
        start_date = datetime.now(timezone.utc).replace(month=1, day=1, hour=0, minute=0, second=0).isoformat()
    
    sales = await db.sales.find({
        "created_at": {"$gte": start_date},
        "status": "completed"
    }, {"_id": 0}).to_list(10000)
    
    revenue = sum(s.get('total', 0) for s in sales)
    
    cost = 0
    for sale in sales:
        for item in sale.get('items', []):
            product = await db.products.find_one({"id": item['product_id']}, {"_id": 0})
            if product:
                cost += product.get('cost_price', 0) * item['quantity']
    
    gross_profit = revenue - cost
    profit_margin = (gross_profit / revenue * 100) if revenue > 0 else 0
    
    return FinancialReport(
        period=period,
        revenue=revenue,
        cost=cost,
        gross_profit=gross_profit,
        profit_margin=profit_margin,
        expenses=0,
        net_profit=gross_profit
    )

# === ADVANCED REPORTS ===

@api_router.get("/reports/abc-curve")
async def get_abc_curve(current_user: dict = Depends(get_current_user)):
    sales = await db.sales.find({"status": "completed"}, {"_id": 0}).to_list(10000)
    product_revenue = {}
    for sale in sales:
        for item in sale.get('items', []):
            pid = item.get('product_id', '')
            pname = item.get('product_name', '')
            rev = item.get('total', 0)
            if pid in product_revenue:
                product_revenue[pid]['revenue'] += rev
                product_revenue[pid]['quantity'] += item.get('quantity', 0)
            else:
                product_revenue[pid] = {'product_id': pid, 'product_name': pname, 'revenue': rev, 'quantity': item.get('quantity', 0)}
    
    items = sorted(product_revenue.values(), key=lambda x: x['revenue'], reverse=True)
    total_rev = sum(i['revenue'] for i in items)
    
    cumulative = 0
    for item in items:
        cumulative += item['revenue']
        pct = (cumulative / total_rev * 100) if total_rev > 0 else 0
        item['percentage'] = round(item['revenue'] / total_rev * 100, 1) if total_rev > 0 else 0
        item['cumulative'] = round(pct, 1)
        if pct <= 80:
            item['class'] = 'A'
        elif pct <= 95:
            item['class'] = 'B'
        else:
            item['class'] = 'C'
    
    return {"items": items, "total_revenue": total_rev}

@api_router.get("/reports/inventory-turnover")
async def get_inventory_turnover(current_user: dict = Depends(get_current_user)):
    products = await db.products.find({"active": True}, {"_id": 0}).to_list(5000)
    sales = await db.sales.find({"status": "completed"}, {"_id": 0}).to_list(10000)
    
    product_sales = {}
    for sale in sales:
        for item in sale.get('items', []):
            pid = item.get('product_id', '')
            qty = item.get('quantity', 0)
            product_sales[pid] = product_sales.get(pid, 0) + qty
    
    inventory = await db.inventory.find({}, {"_id": 0}).to_list(5000)
    product_stock = {}
    for inv in inventory:
        pid = inv.get('product_id', '')
        product_stock[pid] = product_stock.get(pid, 0) + inv.get('quantity', 0)
    
    results = []
    for p in products:
        pid = p['id']
        sold = product_sales.get(pid, 0)
        stock = product_stock.get(pid, 0)
        avg_stock = stock if stock > 0 else 1
        turnover = round(sold / avg_stock, 2) if avg_stock > 0 else 0
        days_cover = round(avg_stock / (sold / 30), 0) if sold > 0 else 999
        
        results.append({
            'product_id': pid,
            'product_name': p['name'],
            'sku': p['sku'],
            'total_sold': sold,
            'current_stock': stock,
            'turnover_rate': turnover,
            'days_of_coverage': min(days_cover, 999),
            'status': 'critico' if days_cover < 7 else 'baixo' if days_cover < 15 else 'normal' if days_cover < 60 else 'excesso'
        })
    
    results.sort(key=lambda x: x['turnover_rate'], reverse=True)
    return {"items": results}

@api_router.get("/audit/export")
async def export_audit_logs(current_user: dict = Depends(get_current_user)):
    await require_role(current_user, ["dev", "master"])
    logs = await db.audit_logs.find({}, {"_id": 0}).sort('timestamp', -1).to_list(10000)
    from report_export import generate_financial_excel
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    import io
    wb = Workbook()
    ws = wb.active
    ws.title = "Auditoria"
    headers = ['Data/Hora', 'Usuario', 'Acao', 'Entidade', 'ID', 'Detalhes']
    hfont = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    hfill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = hfont
        c.fill = hfill
    for i, log in enumerate(logs, 2):
        ws.cell(row=i, column=1, value=log.get('timestamp', ''))
        ws.cell(row=i, column=2, value=log.get('user_email', ''))
        ws.cell(row=i, column=3, value=log.get('action', ''))
        ws.cell(row=i, column=4, value=log.get('entity_type', ''))
        ws.cell(row=i, column=5, value=log.get('entity_id', ''))
        ws.cell(row=i, column=6, value=str(log.get('changes', '') or ''))
    for col in ['A','B','C','D','E','F']:
        ws.column_dimensions[col].width = 22
    buf = io.BytesIO()
    wb.save(buf)
    return Response(content=buf.getvalue(),
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": "attachment; filename=auditoria.xlsx"})

@api_router.post("/invoices/{invoice_id}/process-items")
async def process_invoice_items_to_products(invoice_id: str, current_user: dict = Depends(get_current_user)):
    invoice_doc = await db.invoices.find_one({"id": invoice_id}, {"_id": 0})
    if not invoice_doc:
        raise HTTPException(status_code=404, detail="Nota nao encontrada")
    created = 0
    updated = 0
    for item in invoice_doc.get('items', []):
        pname = item.get('product_name', '')
        if not pname:
            continue
        sku = item.get('product_sku', '') or pname[:10].upper().replace(' ', '')
        qty = item.get('quantity', 0)
        existing = await db.products.find_one({"sku": sku}, {"_id": 0})
        if existing:
            new_avail = existing.get('available_qty', 0) + qty
            await db.products.update_one({"id": existing['id']}, {"$set": {"available_qty": new_avail}})
            updated += 1
        else:
            prod_doc = {
                "id": str(uuid.uuid4()), "name": pname, "sku": sku, "description": "",
                "category": "", "unit": "UN", "min_stock": 0,
                "cost_price": item.get('unit_price', 0), "sale_price": item.get('unit_price', 0),
                "available_qty": qty,
                "active": True, "created_at": datetime.now(timezone.utc).isoformat(),
                "created_by": current_user['user_id']
            }
            await db.products.insert_one(prod_doc)
            created += 1
    await db.invoices.update_one({"id": invoice_id}, {"$set": {"status": "processed"}})
    await audit_logger.log(current_user['user_id'], current_user['email'], "PROCESS", "invoice", invoice_id, {"produtos_criados": created, "produtos_atualizados": updated})
    return {"message": f"Itens enviados para aba Produtos. {created} novos, {updated} atualizados. Transfira para o deposito desejado.", "products_created": created}

@api_router.post("/products/{product_id}/transfer")
async def transfer_product_to_warehouse(product_id: str, warehouse_id: str, quantity: float, sector: str = "", current_user: dict = Depends(get_current_user)):
    product = await db.products.find_one({"id": product_id}, {"_id": 0})
    if not product:
        raise HTTPException(status_code=404, detail="Produto nao encontrado")
    warehouse = await db.warehouses.find_one({"id": warehouse_id}, {"_id": 0})
    if not warehouse:
        raise HTTPException(status_code=404, detail="Deposito nao encontrado")
    
    # Check if product has enough available quantity
    current_stock = product.get('available_qty', quantity)
    
    # Add to inventory
    inv_item = await db.inventory.find_one({"product_id": product_id, "warehouse_id": warehouse_id}, {"_id": 0})
    if inv_item:
        new_qty = inv_item['quantity'] + quantity
        await db.inventory.update_one({"id": inv_item['id']}, {"$set": {"quantity": new_qty, "updated_at": datetime.now(timezone.utc).isoformat()}})
    else:
        await db.inventory.insert_one({
            "id": str(uuid.uuid4()), "product_id": product_id, "warehouse_id": warehouse_id,
            "quantity": quantity, "updated_at": datetime.now(timezone.utc).isoformat()
        })
    
    # Reduce available_qty on product. If zero, delete the product from buffer
    new_available = max(0, current_stock - quantity)
    if new_available <= 0:
        await db.products.delete_one({"id": product_id})
        await audit_logger.log(current_user['user_id'], current_user['email'], "TRANSFER", "product", product_id, {"deposito": warehouse['name'], "setor": sector, "quantidade": quantity, "produto_removido": True})
        return {"message": f"Transferido {quantity} unidades para {warehouse['name']}. Produto removido da aba produtos.", "removed": True}
    else:
        await db.products.update_one({"id": product_id}, {"$set": {"available_qty": new_available}})
        await audit_logger.log(current_user['user_id'], current_user['email'], "TRANSFER", "product", product_id, {"deposito": warehouse['name'], "setor": sector, "quantidade": quantity, "restante": new_available})
        return {"message": f"Transferido {quantity} unidades para {warehouse['name']}. Restam {new_available} na aba produtos.", "removed": False}

@api_router.get("/audit", response_model=List[AuditLog])
async def get_audit_logs(current_user: dict = Depends(get_current_user)):
    await require_role(current_user, ["dev", "master"])
    
    logs = await db.audit_logs.find({}, {"_id": 0}).sort('timestamp', -1).to_list(1000)
    for log in logs:
        log['timestamp'] = datetime.fromisoformat(log['timestamp'])
    return [AuditLog(**log) for log in logs]

# === EDIT ENDPOINTS ===

@api_router.patch("/products/{product_id}", response_model=Product)
async def update_product(product_id: str, updates: ProductUpdate, current_user: dict = Depends(get_current_user)):
    update_data = {k: v for k, v in updates.model_dump().items() if v is not None}
    if not update_data:
        raise HTTPException(status_code=400, detail="No fields to update")
    result = await db.products.update_one({"id": product_id}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Product not found")
    await audit_logger.log(current_user['user_id'], current_user['email'], "UPDATE", "product", product_id, update_data)
    doc = await db.products.find_one({"id": product_id}, {"_id": 0})
    doc['created_at'] = datetime.fromisoformat(doc['created_at'])
    return Product(**doc)

@api_router.delete("/products/{product_id}")
async def delete_product(product_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.products.delete_one({"id": product_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Product not found")
    await audit_logger.log(current_user['user_id'], current_user['email'], "DELETE", "product", product_id)
    return {"message": "Product deleted"}

@api_router.patch("/suppliers/{supplier_id}", response_model=Supplier)
async def update_supplier(supplier_id: str, updates: SupplierUpdate, current_user: dict = Depends(get_current_user)):
    update_data = {k: v for k, v in updates.model_dump().items() if v is not None}
    if not update_data:
        raise HTTPException(status_code=400, detail="No fields to update")
    result = await db.suppliers.update_one({"id": supplier_id}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Supplier not found")
    await audit_logger.log(current_user['user_id'], current_user['email'], "UPDATE", "supplier", supplier_id, update_data)
    doc = await db.suppliers.find_one({"id": supplier_id}, {"_id": 0})
    doc['created_at'] = datetime.fromisoformat(doc['created_at'])
    return Supplier(**doc)

@api_router.delete("/suppliers/{supplier_id}")
async def delete_supplier(supplier_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.suppliers.delete_one({"id": supplier_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Supplier not found")
    await audit_logger.log(current_user['user_id'], current_user['email'], "DELETE", "supplier", supplier_id)
    return {"message": "Supplier deleted"}

@api_router.patch("/warehouses/{warehouse_id}", response_model=Warehouse)
async def update_warehouse(warehouse_id: str, updates: WarehouseUpdate, current_user: dict = Depends(get_current_user)):
    update_data = {k: v for k, v in updates.model_dump().items() if v is not None}
    if not update_data:
        raise HTTPException(status_code=400, detail="No fields to update")
    result = await db.warehouses.update_one({"id": warehouse_id}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Warehouse not found")
    await audit_logger.log(current_user['user_id'], current_user['email'], "UPDATE", "warehouse", warehouse_id, update_data)
    doc = await db.warehouses.find_one({"id": warehouse_id}, {"_id": 0})
    doc['created_at'] = datetime.fromisoformat(doc['created_at'])
    return Warehouse(**doc)

@api_router.delete("/warehouses/{warehouse_id}")
async def delete_warehouse(warehouse_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.warehouses.delete_one({"id": warehouse_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Warehouse not found")
    await audit_logger.log(current_user['user_id'], current_user['email'], "DELETE", "warehouse", warehouse_id)
    return {"message": "Warehouse deleted"}

@api_router.delete("/users/{user_id}")
async def delete_user(user_id: str, current_user: dict = Depends(get_current_user)):
    await require_role(current_user, ["dev"])
    if user_id == current_user['user_id']:
        raise HTTPException(status_code=400, detail="Cannot delete yourself")
    result = await db.users.delete_one({"id": user_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    await audit_logger.log(current_user['user_id'], current_user['email'], "DELETE", "user", user_id)
    return {"message": "User deleted"}

# === ALERT CONFIG ENDPOINTS ===

@api_router.get("/alerts/config", response_model=List[AlertConfig])
async def get_alert_configs(current_user: dict = Depends(get_current_user)):
    configs = await db.alert_configs.find({"user_id": current_user['user_id']}, {"_id": 0}).to_list(100)
    for c in configs:
        c['created_at'] = datetime.fromisoformat(c['created_at'])
    return [AlertConfig(**c) for c in configs]

@api_router.post("/alerts/config", response_model=AlertConfig)
async def create_alert_config(config_data: AlertConfigCreate, current_user: dict = Depends(get_current_user)):
    config = AlertConfig(
        **config_data.model_dump(),
        user_id=current_user['user_id'],
        created_at=datetime.now(timezone.utc)
    )
    doc = config.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.alert_configs.insert_one(doc)
    await audit_logger.log(current_user['user_id'], current_user['email'], "CREATE", "alert_config", config.id)
    return config

@api_router.patch("/alerts/config/{config_id}")
async def update_alert_config(config_id: str, updates: dict, current_user: dict = Depends(get_current_user)):
    updates.pop('id', None)
    updates.pop('user_id', None)
    updates.pop('created_at', None)
    result = await db.alert_configs.update_one(
        {"id": config_id, "user_id": current_user['user_id']},
        {"$set": updates}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Alert config not found")
    return {"message": "Alert config updated"}

@api_router.delete("/alerts/config/{config_id}")
async def delete_alert_config(config_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.alert_configs.delete_one({"id": config_id, "user_id": current_user['user_id']})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Alert config not found")
    return {"message": "Alert config deleted"}

# === NOTIFICATION ENDPOINTS ===

@api_router.get("/notifications")
async def get_notifications(current_user: dict = Depends(get_current_user)):
    notifs = await db.notifications.find(
        {"user_id": current_user['user_id']},
        {"_id": 0}
    ).sort('created_at', -1).to_list(100)
    for n in notifs:
        n['created_at'] = datetime.fromisoformat(n['created_at'])
    return [Notification(**n) for n in notifs]

@api_router.get("/notifications/unread-count")
async def get_unread_count(current_user: dict = Depends(get_current_user)):
    count = await db.notifications.count_documents({"user_id": current_user['user_id'], "read": False})
    return {"count": count}

@api_router.patch("/notifications/{notification_id}/read")
async def mark_notification_read(notification_id: str, current_user: dict = Depends(get_current_user)):
    await db.notifications.update_one(
        {"id": notification_id, "user_id": current_user['user_id']},
        {"$set": {"read": True}}
    )
    return {"message": "Notification marked as read"}

@api_router.post("/notifications/read-all")
async def mark_all_notifications_read(current_user: dict = Depends(get_current_user)):
    await db.notifications.update_many(
        {"user_id": current_user['user_id'], "read": False},
        {"$set": {"read": True}}
    )
    return {"message": "All notifications marked as read"}

@api_router.post("/notifications/send")
async def send_notification(notif_data: NotificationCreate, current_user: dict = Depends(get_current_user)):
    await require_role(current_user, ["dev", "master"])
    notif = Notification(
        **notif_data.model_dump(),
        created_at=datetime.now(timezone.utc)
    )
    doc = notif.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.notifications.insert_one(doc)
    return notif

# === CHECK AND TRIGGER STOCK ALERTS ===

@api_router.post("/alerts/check-stock")
async def check_and_send_stock_alerts(background_tasks: BackgroundTasks, current_user: dict = Depends(get_current_user)):
    await require_role(current_user, ["dev", "master"])
    inventory = await db.inventory.find({}, {"_id": 0}).to_list(5000)
    alerts_sent = 0
    for item in inventory:
        product = await db.products.find_one({"id": item['product_id']}, {"_id": 0})
        warehouse = await db.warehouses.find_one({"id": item['warehouse_id']}, {"_id": 0})
        if product and warehouse and item['quantity'] <= product.get('min_stock', 0):
            configs = await db.alert_configs.find({"alert_type": "stock_low", "active": True}, {"_id": 0}).to_list(100)
            for config in configs:
                if config.get('internal_enabled', True):
                    notif = {
                        "id": str(uuid.uuid4()),
                        "user_id": config['user_id'],
                        "title": "Alerta de Estoque Baixo",
                        "message": f"Produto '{product['name']}' no deposito '{warehouse['name']}' com {item['quantity']} unidades (minimo: {product.get('min_stock', 0)})",
                        "type": "warning",
                        "read": False,
                        "created_at": datetime.now(timezone.utc).isoformat()
                    }
                    await db.notifications.insert_one(notif)
                    alerts_sent += 1
                if config.get('email_enabled') and config.get('email_address'):
                    subject, html = build_stock_alert_email(product['name'], warehouse['name'], item['quantity'], product.get('min_stock', 0))
                    background_tasks.add_task(send_email, config['email_address'], subject, html)
                if config.get('mobile_enabled') and config.get('phone_number'):
                    sms_log = {
                        "id": str(uuid.uuid4()),
                        "phone": config['phone_number'],
                        "message": f"[Gestao TJ] Estoque baixo: {product['name']} ({item['quantity']}/{product.get('min_stock', 0)}) em {warehouse['name']}",
                        "status": "simulated",
                        "created_at": datetime.now(timezone.utc).isoformat()
                    }
                    await db.sms_logs.insert_one(sms_log)
    return {"message": f"{alerts_sent} alerts sent"}

# === FILE UPLOAD FOR INVOICES ===

@api_router.post("/invoices/upload")
async def upload_invoice_file(file: UploadFile = File(...), current_user: dict = Depends(get_current_user)):
    if not file.filename:
        raise HTTPException(status_code=400, detail="No file provided")
    
    content = await file.read()
    filename_lower = file.filename.lower()
    
    if filename_lower.endswith('.xml'):
        try:
            parsed = parse_nfe_xml(content)
            return {"source": "xml", "data": parsed}
        except ValueError as e:
            raise HTTPException(status_code=400, detail=str(e))
    
    elif filename_lower.endswith('.pdf'):
        try:
            api_key = os.environ.get('EMERGENT_LLM_KEY')
            if not api_key:
                raise HTTPException(status_code=500, detail="LLM API key not configured")
            
            b64 = base64.b64encode(content).decode('utf-8')
            
            chat = LlmChat(
                api_key=api_key,
                session_id=f"pdf_{current_user['user_id']}_{datetime.now(timezone.utc).timestamp()}",
                system_message="You are an expert at extracting data from Brazilian fiscal invoices (NFe). Extract all relevant information in JSON format."
            ).with_model("openai", "gpt-4o")
            
            image_content = ImageContent(image_base64=b64)
            
            prompt_text = 'Extract the following information from this Brazilian invoice and return ONLY a valid JSON object with this structure: {"invoice_number": "number", "supplier_name": "name", "supplier_cnpj": "", "issue_date": "YYYY-MM-DD", "total_value": 0.0, "tax_value": 0.0, "items": [{"product_name": "name", "product_sku": "", "quantity": 0.0, "unit_price": 0.0, "total": 0.0, "tax": 0}]}. Return ONLY JSON.'
            
            user_message = UserMessage(text=prompt_text, file_contents=[image_content])
            response = await chat.send_message(user_message)
            
            response_text = response.strip()
            if response_text.startswith('```json'):
                response_text = response_text[7:]
            if response_text.startswith('```'):
                response_text = response_text[3:]
            if response_text.endswith('```'):
                response_text = response_text[:-3]
            
            extracted = json.loads(response_text.strip())
            return {"source": "pdf_ocr", "data": extracted}
        except Exception as e:
            logger.error(f"PDF parsing error: {e}")
            raise HTTPException(status_code=500, detail=f"PDF parsing failed: {str(e)}")
    else:
        raise HTTPException(status_code=400, detail="Unsupported file type. Use PDF or XML.")

# === REPORT EXPORT ENDPOINTS ===

@api_router.get("/reports/export/pdf")
async def export_financial_pdf(period: str, current_user: dict = Depends(get_current_user)):
    # Build report data
    if period == "month":
        start_date = datetime.now(timezone.utc).replace(day=1, hour=0, minute=0, second=0).isoformat()
        period_label = "Mes Atual"
    else:
        start_date = datetime.now(timezone.utc).replace(month=1, day=1, hour=0, minute=0, second=0).isoformat()
        period_label = "Ano Atual"
    
    sales = await db.sales.find({"created_at": {"$gte": start_date}, "status": "completed"}, {"_id": 0}).to_list(10000)
    revenue = sum(s.get('total', 0) for s in sales)
    cost = 0
    for sale in sales:
        for item in sale.get('items', []):
            product = await db.products.find_one({"id": item.get('product_id')}, {"_id": 0})
            if product:
                cost += product.get('cost_price', 0) * item.get('quantity', 0)
    
    gross_profit = revenue - cost
    profit_margin = (gross_profit / revenue * 100) if revenue > 0 else 0
    
    invoices = await db.invoices.find({"created_at": {"$gte": start_date}, "type": "entrada"}, {"_id": 0}).to_list(10000)
    outflows = sum(inv.get('total_value', 0) for inv in invoices)
    
    report_data = {
        "revenue": revenue, "cost": cost, "gross_profit": gross_profit,
        "profit_margin": profit_margin, "expenses": 0, "net_profit": gross_profit,
        "cash_flow": {"inflows": revenue, "outflows": outflows, "balance": revenue - outflows}
    }
    
    pdf_bytes = generate_financial_pdf(report_data, period_label)
    return Response(content=pdf_bytes, media_type="application/pdf",
                    headers={"Content-Disposition": f"attachment; filename=relatorio_{period}.pdf"})

@api_router.get("/reports/export/excel")
async def export_financial_excel(period: str, current_user: dict = Depends(get_current_user)):
    if period == "month":
        start_date = datetime.now(timezone.utc).replace(day=1, hour=0, minute=0, second=0).isoformat()
        period_label = "Mes Atual"
    else:
        start_date = datetime.now(timezone.utc).replace(month=1, day=1, hour=0, minute=0, second=0).isoformat()
        period_label = "Ano Atual"
    
    sales = await db.sales.find({"created_at": {"$gte": start_date}, "status": "completed"}, {"_id": 0}).to_list(10000)
    revenue = sum(s.get('total', 0) for s in sales)
    cost = 0
    for sale in sales:
        for item in sale.get('items', []):
            product = await db.products.find_one({"id": item.get('product_id')}, {"_id": 0})
            if product:
                cost += product.get('cost_price', 0) * item.get('quantity', 0)
    
    gross_profit = revenue - cost
    profit_margin = (gross_profit / revenue * 100) if revenue > 0 else 0
    
    invoices = await db.invoices.find({"created_at": {"$gte": start_date}, "type": "entrada"}, {"_id": 0}).to_list(10000)
    outflows = sum(inv.get('total_value', 0) for inv in invoices)
    
    report_data = {
        "revenue": revenue, "cost": cost, "gross_profit": gross_profit,
        "profit_margin": profit_margin, "expenses": 0, "net_profit": gross_profit,
        "cash_flow": {"inflows": revenue, "outflows": outflows, "balance": revenue - outflows}
    }
    
    excel_bytes = generate_financial_excel(report_data, period_label)
    return Response(content=excel_bytes,
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": f"attachment; filename=relatorio_{period}.xlsx"})

@api_router.post("/seed")
async def seed_database():
    existing_admin = await db.users.find_one({"email": "admin@gestaotj.com"}, {"_id": 0})
    
    if existing_admin:
        return {"message": "Database already seeded"}
    
    users = [
        {
            "id": str(uuid.uuid4()),
            "email": "admin@gestaotj.com",
            "name": "Administrador",
            "role": "dev",
            "active": True,
            "password_hash": hash_password("Admin@123456"),
            "created_at": datetime.now(timezone.utc).isoformat()
        },
        {
            "id": str(uuid.uuid4()),
            "email": "gerente@gestaotj.com",
            "name": "Gerente",
            "role": "master",
            "active": True,
            "password_hash": hash_password("Gerente@123"),
            "created_at": datetime.now(timezone.utc).isoformat()
        },
        {
            "id": str(uuid.uuid4()),
            "email": "usuario@gestaotj.com",
            "name": "Usuário",
            "role": "usuario",
            "active": True,
            "password_hash": hash_password("Usuario@123"),
            "created_at": datetime.now(timezone.utc).isoformat()
        }
    ]
    
    await db.users.insert_many(users)
    
    return {"message": "Database seeded successfully"}

app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
