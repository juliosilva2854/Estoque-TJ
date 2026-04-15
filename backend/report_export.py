import io
from datetime import datetime
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def generate_financial_pdf(report_data: dict, period_label: str) -> bytes:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=20*mm, bottomMargin=20*mm)
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle('CustomTitle', parent=styles['Title'], fontSize=22, textColor=colors.HexColor('#09090B'))
    subtitle_style = ParagraphStyle('CustomSubtitle', parent=styles['Normal'], fontSize=12, textColor=colors.HexColor('#71717A'))
    
    elements = []
    
    elements.append(Paragraph("Gestao TJ - Relatorio Financeiro", title_style))
    elements.append(Spacer(1, 5*mm))
    elements.append(Paragraph(f"Periodo: {period_label} | Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}", subtitle_style))
    elements.append(Spacer(1, 10*mm))
    
    # DRE Table
    dre_data = [
        ['Descricao', 'Valor (R$)'],
        ['Receita Bruta', f"{report_data.get('revenue', 0):,.2f}"],
        ['(-) Custo dos Produtos', f"({report_data.get('cost', 0):,.2f})"],
        ['(=) Lucro Bruto', f"{report_data.get('gross_profit', 0):,.2f}"],
        ['(-) Despesas Operacionais', f"({report_data.get('expenses', 0):,.2f})"],
        ['(=) Lucro Liquido', f"{report_data.get('net_profit', 0):,.2f}"],
        ['Margem de Lucro', f"{report_data.get('profit_margin', 0):.1f}%"],
    ]
    
    table = Table(dre_data, colWidths=[300, 200])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2563EB')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#FAFAFA')]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E4E4E7')),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('LEFTPADDING', (0, 0), (-1, -1), 12),
        ('RIGHTPADDING', (0, 0), (-1, -1), 12),
        ('BACKGROUND', (0, -2), (-1, -2), colors.HexColor('#F0FDF4')),
        ('FONTNAME', (0, -2), (-1, -2), 'Helvetica-Bold'),
    ]))
    elements.append(table)
    
    # Cash Flow section
    if 'cash_flow' in report_data:
        elements.append(Spacer(1, 15*mm))
        elements.append(Paragraph("Fluxo de Caixa", title_style))
        elements.append(Spacer(1, 5*mm))
        
        cf = report_data['cash_flow']
        cf_data = [
            ['Descricao', 'Valor (R$)'],
            ['Entradas (Vendas)', f"{cf.get('inflows', 0):,.2f}"],
            ['Saidas (Compras/NFs)', f"({cf.get('outflows', 0):,.2f})"],
            ['Saldo do Periodo', f"{cf.get('balance', 0):,.2f}"],
        ]
        
        cf_table = Table(cf_data, colWidths=[300, 200])
        cf_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#16A34A')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E4E4E7')),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('LEFTPADDING', (0, 0), (-1, -1), 12),
            ('RIGHTPADDING', (0, 0), (-1, -1), 12),
        ]))
        elements.append(cf_table)
    
    doc.build(elements)
    return buffer.getvalue()


def generate_financial_excel(report_data: dict, period_label: str) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "DRE"
    
    header_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
    money_font = Font(name='Calibri', size=11)
    bold_font = Font(name='Calibri', size=11, bold=True)
    green_fill = PatternFill(start_color='DCFCE7', end_color='DCFCE7', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin', color='E4E4E7'),
        right=Side(style='thin', color='E4E4E7'),
        top=Side(style='thin', color='E4E4E7'),
        bottom=Side(style='thin', color='E4E4E7'),
    )
    
    # Title
    ws.merge_cells('A1:B1')
    ws['A1'] = f"Gestao TJ - Relatorio Financeiro ({period_label})"
    ws['A1'].font = Font(name='Calibri', size=16, bold=True)
    
    ws.merge_cells('A2:B2')
    ws['A2'] = f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws['A2'].font = Font(name='Calibri', size=10, color='71717A')
    
    # DRE Header
    row = 4
    for col_idx, header in enumerate(['Descricao', 'Valor (R$)'], 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    # DRE Data
    dre_rows = [
        ('Receita Bruta', report_data.get('revenue', 0)),
        ('(-) Custo dos Produtos', -report_data.get('cost', 0)),
        ('(=) Lucro Bruto', report_data.get('gross_profit', 0)),
        ('(-) Despesas Operacionais', -report_data.get('expenses', 0)),
        ('(=) Lucro Liquido', report_data.get('net_profit', 0)),
        ('Margem de Lucro (%)', report_data.get('profit_margin', 0)),
    ]
    
    for desc, val in dre_rows:
        row += 1
        ws.cell(row=row, column=1, value=desc).font = bold_font if '(=)' in desc else money_font
        ws.cell(row=row, column=1).border = thin_border
        cell = ws.cell(row=row, column=2, value=val)
        cell.font = bold_font if '(=)' in desc else money_font
        cell.number_format = '#,##0.00' if 'Margem' not in desc else '0.0"%"'
        cell.alignment = Alignment(horizontal='right')
        cell.border = thin_border
        if 'Lucro Liquido' in desc:
            ws.cell(row=row, column=1).fill = green_fill
            cell.fill = green_fill
    
    # Cash Flow Sheet
    if 'cash_flow' in report_data:
        ws2 = wb.create_sheet("Fluxo de Caixa")
        ws2.merge_cells('A1:B1')
        ws2['A1'] = "Fluxo de Caixa"
        ws2['A1'].font = Font(name='Calibri', size=16, bold=True)
        
        cf = report_data['cash_flow']
        cf_row = 3
        for col_idx, header in enumerate(['Descricao', 'Valor (R$)'], 1):
            cell = ws2.cell(row=cf_row, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = PatternFill(start_color='16A34A', end_color='16A34A', fill_type='solid')
            cell.border = thin_border
        
        for desc, val in [
            ('Entradas (Vendas)', cf.get('inflows', 0)),
            ('Saidas (Compras/NFs)', cf.get('outflows', 0)),
            ('Saldo do Periodo', cf.get('balance', 0)),
        ]:
            cf_row += 1
            ws2.cell(row=cf_row, column=1, value=desc).border = thin_border
            cell = ws2.cell(row=cf_row, column=2, value=val)
            cell.number_format = '#,##0.00'
            cell.alignment = Alignment(horizontal='right')
            cell.border = thin_border
        
        ws2.column_dimensions['A'].width = 30
        ws2.column_dimensions['B'].width = 20
    
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20
    
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
